from __future__ import annotations

import argparse
import json
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import AppPaths
from .db import connect, init_db


HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")


CATEGORY_RULES: list[dict[str, object]] = [
    {
        "name": "채널/인터뷰/특집",
        "strong": (
            "슈카의 변",
            "초대석",
            "인터뷰",
            "공지",
            "사과",
            "특집",
            "모음",
            "생방",
            "라이브",
        ),
        "medium": ("채널", "방송", "복귀", "윤하", "게스트"),
    },
    {
        "name": "국제정세/전쟁",
        "strong": (
            "전쟁",
            "휴전",
            "침공",
            "정상회담",
            "외교",
            "군사",
            "핵",
            "우크라",
            "러시아",
            "미국",
            "중국",
            "북한",
            "이란",
            "대만",
        ),
        "medium": ("일본", "유럽", "트럼프", "대통령", "민주주의", "동맹", "제재"),
    },
    {
        "name": "경제/금융/투자",
        "strong": (
            "금리",
            "주식",
            "증시",
            "투자",
            "채권",
            "환율",
            "인플레이션",
            "인플레",
            "은행",
            "연금",
            "리만",
            "게임스탑",
            "버핏",
            "코인",
            "비트코인",
            "파산",
            "폭락",
        ),
        "medium": ("경기", "재정", "부채", "적자", "흑자", "예금", "금융", "펀드"),
    },
    {
        "name": "산업/기업/기술",
        "strong": (
            "ai",
            "반도체",
            "엔비디아",
            "테슬라",
            "머스크",
            "배터리",
            "자동차",
            "삼성",
            "애플",
            "카카오",
            "오픈ai",
            "스타트업",
        ),
        "medium": ("기업", "플랫폼", "유튜브", "넷플릭스", "로봇", "기술", "공장", "칩"),
    },
    {
        "name": "사회/정책",
        "strong": (
            "부동산",
            "전세",
            "출산",
            "인구",
            "노동",
            "복지",
            "교육",
            "의료",
            "건강보험",
            "세금",
            "규제",
            "정책",
            "법",
            "저출산",
        ),
        "medium": ("청년", "연봉", "임금", "집값", "집", "대한민국", "한국"),
    },
    {
        "name": "역사/인물",
        "strong": (
            "역사",
            "제국",
            "왕",
            "황제",
            "로마",
            "조선",
            "징기스칸",
            "나폴레옹",
            "푸틴",
            "히틀러",
            "전설",
            "혁명",
            "학살",
        ),
        "medium": ("중세", "고대", "왕실", "장군", "인물", "비극"),
    },
    {
        "name": "문화/콘텐츠/생활",
        "strong": (
            "영화",
            "드라마",
            "k-pop",
            "콘텐츠",
            "월드컵",
            "여행",
            "결혼",
            "빵",
            "음식",
            "가수",
            "데몬헌터스",
            "ces",
        ),
        "medium": ("문화", "생활", "노래", "책", "게임", "애니", "축구"),
    },
    {
        "name": "과학/환경/에너지",
        "strong": (
            "nasa",
            "우주",
            "기후",
            "과학",
            "환경",
            "에너지",
            "원유",
            "석유",
            "가스",
            "원전",
            "지구",
            "기상이변",
        ),
        "medium": ("온난화", "태양", "전기", "행성", "자원", "생태", "발전"),
    },
    {
        "name": "사건/범죄/재난",
        "strong": (
            "횡령",
            "사기",
            "참사",
            "재난",
            "살인",
            "담합",
            "루머",
            "대지진",
            "사망",
            "붕괴",
            "최악",
        ),
        "medium": ("충격", "공포", "위기", "사건", "비극", "멸망"),
    },
]


def parse_keywords_text(raw: str | None) -> str:
    if not raw:
        return ""
    try:
        value = json.loads(raw)
    except (TypeError, ValueError):
        return str(raw)
    if isinstance(value, list):
        return " ".join(str(item) for item in value if item)
    if isinstance(value, dict):
        return " ".join(str(item) for item in value.values() if item)
    return str(value)


def score_matches(text: str, terms: Iterable[str], weight: int) -> int:
    return sum(weight for term in terms if term and term.lower() in text)


def categorize_video(title: str, summary: str = "", keywords_json: str = "") -> str:
    title_text = (title or "").strip().lower()
    aux_text = " ".join(
        part.strip().lower() for part in (summary or "", parse_keywords_text(keywords_json)) if part
    )
    full_text = " ".join(part for part in (title_text, aux_text) if part)
    if not full_text:
        return "기타/일반"

    if any(term in title_text for term in ("초대석", "인터뷰", "특집", "모음", "슈카의 변")):
        return "채널/인터뷰/특집"
    if any(term in full_text for term in ("전쟁", "침공", "휴전")) and any(
        term in full_text for term in ("미국", "중국", "러시아", "우크라", "북한", "이란", "대만")
    ):
        return "국제정세/전쟁"
    if any(term in full_text for term in ("주식", "금리", "환율", "채권", "연금", "투자")):
        return "경제/금융/투자"
    if any(term in full_text for term in ("반도체", "엔비디아", "ai", "테슬라", "오픈ai")):
        return "산업/기업/기술"

    scored: list[tuple[int, str]] = []
    for rule in CATEGORY_RULES:
        name = str(rule["name"])
        strong = tuple(str(term) for term in rule.get("strong", ()))
        medium = tuple(str(term) for term in rule.get("medium", ()))
        score = 0
        score += score_matches(title_text, strong, 6)
        score += score_matches(aux_text, strong, 3)
        score += score_matches(title_text, medium, 3)
        score += score_matches(aux_text, medium, 1)

        if name == "국제정세/전쟁" and any(term in full_text for term in ("미국", "중국", "일본", "러시아", "우크라", "북한", "이란", "대만")):
            score += 3
        if name == "경제/금융/투자" and any(term in full_text for term in ("시장", "경제", "금융", "증시", "은행", "투자")):
            score += 3
        if name == "산업/기업/기술" and any(term in full_text for term in ("기업", "기술", "산업", "플랫폼")):
            score += 2
        if name == "사회/정책" and any(term in full_text for term in ("정책", "규제", "복지", "노동")):
            score += 2
        if name == "역사/인물" and any(term in full_text for term in ("역사", "제국", "왕", "혁명", "전설")):
            score += 3
        if name == "문화/콘텐츠/생활" and any(term in full_text for term in ("영화", "가수", "콘텐츠", "여행", "생활")):
            score += 2
        if name == "과학/환경/에너지" and any(term in full_text for term in ("우주", "에너지", "환경", "기후")):
            score += 2
        if name == "사건/범죄/재난" and any(term in full_text for term in ("사건", "재난", "범죄", "횡령", "참사")):
            score += 3

        scored.append((score, name))

    scored.sort(key=lambda item: item[0], reverse=True)
    if scored and scored[0][0] >= 4:
        return scored[0][1]
    return "기타/일반"


def latest_attempts(conn, stage: str) -> dict[str, dict]:
    rows = conn.execute(
        """
        SELECT video_id, status, attempts, returncode, stderr, created_at
        FROM (
            SELECT
                video_id,
                status,
                attempts,
                returncode,
                stderr,
                created_at,
                ROW_NUMBER() OVER (PARTITION BY video_id ORDER BY id DESC) AS rn
            FROM download_attempts
            WHERE stage = ?
        )
        WHERE rn = 1
        """,
        (stage,),
    ).fetchall()
    return {
        row["video_id"]: {
            "status": row["status"],
            "attempts": row["attempts"],
            "returncode": row["returncode"],
            "stderr": row["stderr"] or "",
            "created_at": row["created_at"],
        }
        for row in rows
    }


def normalize_reason(stderr: str) -> str:
    text = (stderr or "").strip()
    if not text:
        return ""
    text_lower = text.lower()
    if "sign in to confirm your age" in text_lower:
        return "Age restricted"
    if "requested format is not available" in text_lower:
        return "Requested format unavailable"
    if "offline." in text_lower:
        return "Offline stream/video"
    if "this live event will begin in a few moments" in text_lower:
        return "Upcoming live stream"
    if "premieres in" in text_lower:
        return "Upcoming premiere"
    if "ffmpeg not found" in text_lower:
        return "ffmpeg missing"
    return text.splitlines()[0][:120]


def latest_issue(stage_attempt: dict | None) -> str:
    if not stage_attempt:
        return ""
    return normalize_reason(stage_attempt.get("stderr", ""))


def derive_collection_status(video: dict, info_attempt: dict | None, subtitle_attempt: dict | None) -> str:
    has_info_json = bool(video["has_info_json"])
    has_transcript = bool(video["has_transcript"])
    has_ko_sub = bool(video["has_ko_sub"])
    has_auto_ko_sub = bool(video["has_auto_ko_sub"])

    if has_transcript:
        return "completed"
    if info_attempt and info_attempt["status"] == "skipped":
        return "info_json_skipped"
    if not has_info_json:
        return "info_json_missing"
    if not has_ko_sub and not has_auto_ko_sub:
        return "metadata_only"
    if subtitle_attempt and subtitle_attempt["status"] == "skipped":
        return "subtitle_skipped"
    if subtitle_attempt and subtitle_attempt["status"] == "failed":
        return "subtitle_failed"
    return "subtitle_pending"


def thumbnail_exists(paths: AppPaths, video_id: str) -> bool:
    return any(paths.thumbnails_dir.glob(f"*_{video_id}_*"))


def auto_fit(ws) -> None:
    for column_cells in ws.columns:
        length = 0
        column = column_cells[0].column
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, len(value))
        ws.column_dimensions[get_column_letter(column)].width = min(max(length + 2, 10), 48)


def style_header(ws, row_idx: int = 1) -> None:
    for cell in ws[row_idx]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def append_table(ws, headers: list[str], rows: Iterable[Iterable]) -> None:
    ws.append(headers)
    style_header(ws, ws.max_row)
    for row in rows:
        ws.append(list(row))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    auto_fit(ws)


def build_summary_sheet(wb: Workbook, stats: dict, status_counts: Counter, info_issue_counts: Counter, subtitle_issue_counts: Counter) -> None:
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Metric", "Value"])
    style_header(ws)
    rows = [
        ("Generated at", stats["generated_at"]),
        ("Videos", stats["videos"]),
        ("Info JSON linked", stats["info_json"]),
        ("Info JSON missing", stats["info_json_missing"]),
        ("Videos with manual ko_sub", stats["ko_sub"]),
        ("Videos with auto ko_sub", stats["auto_ko_sub"]),
        ("Videos with any ko_sub", stats["any_ko_sub"]),
        ("Transcripts", stats["transcripts"]),
        ("Thumbnail files", stats["thumbnails"]),
    ]
    for row in rows:
        ws.append(row)

    ws.append([])
    ws.append(["Collection status", "Count"])
    style_header(ws, ws.max_row)
    for status, count in sorted(status_counts.items()):
        ws.append((status, count))

    ws.append([])
    ws.append(["Info JSON issue", "Count"])
    style_header(ws, ws.max_row)
    for reason, count in info_issue_counts.most_common():
        ws.append((reason, count))

    ws.append([])
    ws.append(["Subtitle issue", "Count"])
    style_header(ws, ws.max_row)
    for reason, count in subtitle_issue_counts.most_common():
        ws.append((reason, count))

    auto_fit(ws)


def build_yearly_sheet(wb: Workbook, conn) -> None:
    ws = wb.create_sheet("Yearly")
    rows = conn.execute(
        """
        SELECT
            substr(upload_date, 1, 4) AS year,
            COUNT(*) AS videos,
            SUM(CASE WHEN info_json_path IS NOT NULL AND info_json_path != '' THEN 1 ELSE 0 END) AS info_json,
            SUM(has_ko_sub) AS ko_sub,
            SUM(has_auto_ko_sub) AS auto_ko_sub,
            SUM(CASE WHEN has_ko_sub = 1 OR has_auto_ko_sub = 1 THEN 1 ELSE 0 END) AS any_ko_sub,
            SUM(CASE WHEN video_id IN (SELECT video_id FROM transcripts) THEN 1 ELSE 0 END) AS transcripts
        FROM videos
        GROUP BY substr(upload_date, 1, 4)
        ORDER BY year
        """
    ).fetchall()
    append_table(
        ws,
        ["Year", "Videos", "Info JSON", "Info JSON %", "Manual ko_sub", "Auto ko_sub", "Any ko_sub", "Transcripts", "Transcript % of any ko_sub"],
        (
            (
                row["year"] or "(blank)",
                row["videos"],
                row["info_json"],
                round((row["info_json"] or 0) / row["videos"], 4) if row["videos"] else 0,
                row["ko_sub"] or 0,
                row["auto_ko_sub"] or 0,
                row["any_ko_sub"] or 0,
                row["transcripts"] or 0,
                round((row["transcripts"] or 0) / ((row["any_ko_sub"] or 0) or 1), 4)
                if (row["any_ko_sub"] or 0)
                else 0,
            )
            for row in rows
        ),
    )


def build_videos_sheet(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("Videos")
    headers = [
        "upload_date",
        "video_id",
        "category",
        "title",
        "view_count",
        "like_count",
        "like_ratio",
        "collection_status",
        "has_info_json",
        "has_thumbnail",
        "has_ko_sub",
        "has_auto_ko_sub",
        "has_transcript",
        "transcript_source",
        "info_json_status",
        "info_json_reason",
        "subtitle_status",
        "subtitle_reason",
        "segment_count",
        "source_url",
    ]
    append_table(
        ws,
        headers,
        (
            (
                row["upload_date"],
                row["video_id"],
                row["category"],
                row["title"],
                row["view_count"],
                row["like_count"],
                row["like_ratio"],
                row["collection_status"],
                row["has_info_json"],
                row["has_thumbnail"],
                row["has_ko_sub"],
                row["has_auto_ko_sub"],
                row["has_transcript"],
                row["transcript_source"],
                row["info_json_status"],
                row["info_json_reason"],
                row["subtitle_status"],
                row["subtitle_reason"],
                row["segment_count"],
                row["source_url"],
            )
            for row in rows
        ),
    )


def build_category_sheet(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("Categories")
    counts = Counter(row["category"] for row in rows)
    append_table(
        ws,
        ["category", "videos", "avg_views", "median_views", "avg_like_ratio"],
        (
            (
                category,
                len(category_rows),
                round(sum((row["view_count"] or 0) for row in category_rows) / len(category_rows), 1),
                sorted((row["view_count"] or 0) for row in category_rows)[len(category_rows) // 2],
                round(
                    sum((row["like_ratio"] or 0) for row in category_rows if row["like_ratio"] is not None)
                    / max(sum(1 for row in category_rows if row["like_ratio"] is not None), 1),
                    4,
                ),
            )
            for category, category_rows in sorted(
                (
                    (category, [row for row in rows if row["category"] == category])
                    for category in counts
                ),
                key=lambda item: (-len(item[1]), item[0]),
            )
        ),
    )


def build_issues_sheet(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("Issues")
    issue_rows = [
        row
        for row in rows
        if row["collection_status"] in {"info_json_missing", "info_json_skipped", "subtitle_pending", "subtitle_failed", "subtitle_skipped"}
    ]
    append_table(
        ws,
        ["upload_date", "video_id", "title", "collection_status", "info_json_reason", "subtitle_reason", "source_url"],
        (
            (
                row["upload_date"],
                row["video_id"],
                row["title"],
                row["collection_status"],
                row["info_json_reason"],
                row["subtitle_reason"],
                row["source_url"],
            )
            for row in issue_rows
        ),
    )


def build_attempts_sheet(wb: Workbook, conn) -> None:
    ws = wb.create_sheet("LatestAttempts")
    rows = conn.execute(
        """
        SELECT video_id, stage, status, attempts, returncode, created_at, substr(coalesce(stderr, ''), 1, 180) AS stderr_short
        FROM (
            SELECT
                video_id,
                stage,
                status,
                attempts,
                returncode,
                created_at,
                stderr,
                ROW_NUMBER() OVER (PARTITION BY video_id, stage ORDER BY id DESC) AS rn
            FROM download_attempts
        )
        WHERE rn = 1
        ORDER BY created_at DESC
        """
    ).fetchall()
    append_table(
        ws,
        ["video_id", "stage", "status", "attempts", "returncode", "created_at", "stderr_short"],
        (
            (
                row["video_id"],
                row["stage"],
                row["status"],
                row["attempts"],
                row["returncode"],
                row["created_at"],
                row["stderr_short"],
            )
            for row in rows
        ),
    )


def generate_report(base_dir: str, output_path: str | None = None) -> Path:
    paths = AppPaths.from_base_dir(base_dir)
    conn = connect(paths.db_path)
    init_db(conn)
    info_attempts = latest_attempts(conn, "info_json")
    subtitle_attempts = latest_attempts(conn, "subtitle")

    video_rows = conn.execute(
        """
        SELECT
            v.video_id,
            v.title,
            v.upload_date,
            v.view_count,
            v.like_count,
            v.has_ko_sub,
            v.has_auto_ko_sub,
            v.source_url,
            v.info_json_path,
            CASE WHEN t.video_id IS NULL THEN 0 ELSE 1 END AS has_transcript,
            t.subtitle_source,
            COALESCE(t.segment_count, 0) AS segment_count,
            a.summary,
            a.keywords_json
        FROM videos v
        LEFT JOIN transcripts t ON t.video_id = v.video_id
        LEFT JOIN video_analysis a ON a.video_id = v.video_id
        ORDER BY v.upload_date DESC, v.video_id DESC
        """
    ).fetchall()

    prepared_rows: list[dict] = []
    status_counts: Counter = Counter()
    info_issue_counts: Counter = Counter()
    subtitle_issue_counts: Counter = Counter()

    for row in video_rows:
        info_attempt = info_attempts.get(row["video_id"])
        subtitle_attempt = subtitle_attempts.get(row["video_id"])
        prepared = {
            "video_id": row["video_id"],
            "title": row["title"],
            "upload_date": row["upload_date"],
            "view_count": row["view_count"],
            "like_count": row["like_count"],
            "source_url": row["source_url"],
            "has_info_json": bool(row["info_json_path"]),
            "has_thumbnail": thumbnail_exists(paths, row["video_id"]),
            "has_ko_sub": bool(row["has_ko_sub"]),
            "has_auto_ko_sub": bool(row["has_auto_ko_sub"]),
            "has_transcript": bool(row["has_transcript"]),
            "transcript_source": row["subtitle_source"] or "",
            "segment_count": row["segment_count"],
            "info_json_status": info_attempt["status"] if info_attempt else "",
            "info_json_reason": latest_issue(info_attempt),
            "subtitle_status": subtitle_attempt["status"] if subtitle_attempt else "",
            "subtitle_reason": latest_issue(subtitle_attempt),
            "category": categorize_video(
                row["title"],
                row["summary"] or "",
                row["keywords_json"] or "",
            ),
        }
        if row["view_count"] and row["like_count"] is not None:
            prepared["like_ratio"] = round(row["like_count"] / row["view_count"], 4)
        else:
            prepared["like_ratio"] = None
        prepared["collection_status"] = derive_collection_status(prepared, info_attempt, subtitle_attempt)
        status_counts[prepared["collection_status"]] += 1
        if prepared["info_json_reason"]:
            info_issue_counts[prepared["info_json_reason"]] += 1
        if prepared["subtitle_reason"]:
            subtitle_issue_counts[prepared["subtitle_reason"]] += 1
        prepared_rows.append(prepared)

    thumbnail_count = sum(1 for _ in paths.thumbnails_dir.glob("*.jpg"))
    stats = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "videos": len(prepared_rows),
        "info_json": sum(1 for row in prepared_rows if row["has_info_json"]),
        "info_json_missing": sum(1 for row in prepared_rows if not row["has_info_json"]),
        "ko_sub": sum(1 for row in prepared_rows if row["has_ko_sub"]),
        "auto_ko_sub": sum(1 for row in prepared_rows if row["has_auto_ko_sub"]),
        "any_ko_sub": sum(1 for row in prepared_rows if row["has_ko_sub"] or row["has_auto_ko_sub"]),
        "transcripts": sum(1 for row in prepared_rows if row["has_transcript"]),
        "thumbnails": thumbnail_count,
    }

    wb = Workbook()
    build_summary_sheet(wb, stats, status_counts, info_issue_counts, subtitle_issue_counts)
    build_yearly_sheet(wb, conn)
    build_videos_sheet(wb, prepared_rows)
    build_category_sheet(wb, prepared_rows)
    build_issues_sheet(wb, prepared_rows)
    build_attempts_sheet(wb, conn)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    reports_dir = paths.base_dir / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)
    if output_path:
        output = Path(output_path).resolve()
        output.parent.mkdir(parents=True, exist_ok=True)
    else:
        output = reports_dir / f"syuka_collection_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(output)
    conn.close()
    return output


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Generate Syuka Ops collection report")
    parser.add_argument("--base-dir", default="./data")
    parser.add_argument("--output", default=None, help="Output .xlsx path")
    return parser


def main() -> None:
    args = build_parser().parse_args()
    output = generate_report(args.base_dir, args.output)
    print(output)


if __name__ == "__main__":
    main()
